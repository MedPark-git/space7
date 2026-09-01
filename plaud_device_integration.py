import hashlib
import hmac
import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from urllib.parse import urlsplit
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import portal_core as core


logger = logging.getLogger("medpark.plaud_device")
KST = ZoneInfo("Asia/Seoul")
WEBHOOK_PATH = "/api/integrations/plaud-device/webhook"
_memory_meetings = {}


def _webhook_secret():
    return (os.environ.get("PLAUD_DEVICE_WEBHOOK_SECRET") or "").strip()


def configured():
    return len(_webhook_secret()) >= 16


def configuration_payload():
    ready = configured()
    return {
        "provider": "PLAUD Note Pro + Zapier",
        "configured": ready,
        "status": "ready" if ready else "waiting",
        "message": (
            "Zapier 회의록 수신 준비가 완료되었습니다."
            if ready else
            "PLAUD_DEVICE_WEBHOOK_SECRET 환경변수를 등록하면 Zapier 수신을 시작할 수 있습니다."
        ),
        "webhook_path": WEBHOOK_PATH,
        "auth_header": "X-MedPark-Webhook-Secret",
        "device_model": "PN0300",
        "title_masking": "default",
        "excel_export": True,
    }


def _clean_text(value, maximum=20000):
    text = str(value or "").replace("\x00", "").strip()
    return text[:maximum]


def _clean_title(value):
    title = re.sub(r"\s+", " ", _clean_text(value, 200))
    return title or "PLAUD 기기 회의록"


def mask_title(title):
    text = _clean_title(title)
    if len(text) <= 1:
        return "•"
    visible = min(2, len(text))
    return f"{text[:visible]}{'•' * max(4, min(12, len(text) - visible))}"


def _bounded_int(value, default, minimum, maximum):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(minimum, min(maximum, parsed))


def _duration(value):
    try:
        return max(0.0, min(float(value or 0), 24 * 60 * 60))
    except (TypeError, ValueError):
        return 0.0


def _parse_datetime(value):
    if isinstance(value, datetime):
        parsed = value
    else:
        raw = _clean_text(value, 80)
        if not raw:
            return datetime.now(timezone.utc)
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError as exc:
            raise core.AppError("회의 일시는 ISO 8601 형식으로 전송해 주세요.") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=KST)
    return parsed.astimezone(timezone.utc)


def _text_list(value, maximum_items=100):
    if value is None:
        return []
    if isinstance(value, str):
        candidates = re.split(r"[\r\n,]+", value)
    elif isinstance(value, (list, tuple)):
        candidates = list(value)
    else:
        candidates = [value]
    normalized = []
    for item in candidates[:maximum_items]:
        if isinstance(item, dict):
            item = item.get("text") or item.get("task") or item.get("title") or item.get("content") or json.dumps(item, ensure_ascii=False)
        text = _clean_text(item, 500)
        if text:
            normalized.append(text)
    return normalized


def _source_url(value):
    raw = _clean_text(value, 2000)
    if not raw:
        return None
    parsed = urlsplit(raw)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise core.AppError("PLAUD 원본 링크는 http/https 주소만 사용할 수 있습니다.")
    return raw


def _status(value, transcript):
    status = _clean_text(value, 20).lower()
    if status in {"completed", "processing", "failed"}:
        return status
    return "completed" if transcript else "processing"


def _external_id(data, title, meeting_date, transcript):
    candidate = data.get("external_id") or data.get("recording_id") or data.get("note_id") or data.get("id")
    candidate = _clean_text(candidate, 255)
    if candidate:
        return candidate
    digest = hashlib.sha256(
        f"{title}|{meeting_date.isoformat()}|{transcript[:1000]}".encode("utf-8")
    ).hexdigest()
    return f"generated-{digest}"


def _normalize_payload(data):
    if not isinstance(data, dict):
        raise core.AppError("Zapier 회의록 데이터가 올바르지 않습니다.")
    title = _clean_title(data.get("title") or data.get("meeting_title") or data.get("name"))
    transcript = _clean_text(data.get("transcript") or data.get("content"), 500000)
    meeting_date = _parse_datetime(data.get("meeting_date") or data.get("created_at") or data.get("recorded_at"))
    return {
        "external_id": _external_id(data, title, meeting_date, transcript),
        "source": "plaud_note_pro",
        "device_model": _clean_text(data.get("device_model") or "PN0300", 50),
        "title": title,
        "meeting_date": meeting_date,
        "duration_seconds": _duration(data.get("duration_seconds") or data.get("duration")),
        "participants": _text_list(data.get("participants") or data.get("attendees")),
        "transcript": transcript,
        "summary": _clean_text(data.get("summary"), 100000),
        "decisions": _text_list(data.get("decisions")),
        "action_items": _text_list(data.get("action_items") or data.get("tasks")),
        "status": _status(data.get("status"), transcript),
        "source_url": _source_url(data.get("source_url") or data.get("url")),
    }


def _authorized(provided_secret):
    expected = _webhook_secret()
    return bool(expected) and hmac.compare_digest(expected, str(provided_secret or "").strip())


def accept_webhook(data, provided_secret, ip=None):
    if not configured():
        raise core.AppError("Zapier 수신 보안키가 아직 등록되지 않았습니다.", 503)
    if not _authorized(provided_secret):
        logger.warning("PLAUD device webhook rejected ip=%s", ip or "unknown")
        raise core.AppError("Zapier 웹훅 인증에 실패했습니다.", 401)
    values = _normalize_payload(data)
    meeting_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    if core.DB_ENABLED:
        row = core.execute(
            """INSERT INTO plaud_device_meetings(
                 id,external_id,source,device_model,title,meeting_date,duration_seconds,
                 participants,transcript,summary,decisions,action_items,status,source_url
               ) VALUES(%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s,%s,%s::jsonb,%s::jsonb,%s,%s)
               ON CONFLICT(external_id) DO UPDATE SET
                 source=excluded.source,device_model=excluded.device_model,title=excluded.title,
                 meeting_date=excluded.meeting_date,duration_seconds=excluded.duration_seconds,
                 participants=excluded.participants,transcript=excluded.transcript,summary=excluded.summary,
                 decisions=excluded.decisions,action_items=excluded.action_items,status=excluded.status,
                 source_url=excluded.source_url,updated_at=now()
               RETURNING *""",
            (
                meeting_id, values["external_id"], values["source"], values["device_model"], values["title"],
                values["meeting_date"], values["duration_seconds"], json.dumps(values["participants"], ensure_ascii=False),
                values["transcript"], values["summary"], json.dumps(values["decisions"], ensure_ascii=False),
                json.dumps(values["action_items"], ensure_ascii=False), values["status"], values["source_url"],
            ),
        )
    else:
        existing = next((item for item in _memory_meetings.values() if item["external_id"] == values["external_id"]), None)
        if existing:
            existing.update(values)
            existing["updated_at"] = now
            row = existing
        else:
            row = {"id": meeting_id, "received_at": now, "updated_at": now, **values}
            _memory_meetings[meeting_id] = row
    try:
        core.write_audit(None, "plaud_device.meeting.receive", "plaud_device_meeting", str(row["id"]), {"external_id": values["external_id"]}, ip)
    except Exception:
        logger.exception("PLAUD device audit write failed meeting_id=%s", str(row.get("id")))
    logger.info("PLAUD device meeting received meeting_id=%s external_id=%s", str(row["id"]), values["external_id"])
    return {"success": True, "meeting_id": str(row["id"]), "external_id": values["external_id"], "status": row.get("status")}


def _can_reveal(actor, requested=False):
    return bool(requested) and bool(actor) and actor.get("role") == "admin"


def public_meeting(row, actor, detail=False, reveal_titles=False):
    reveal = _can_reveal(actor, reveal_titles)
    result = {
        "id": str(row.get("id")),
        "title": row.get("title") if reveal else None,
        "masked_title": mask_title(row.get("title")),
        "title_revealed": reveal,
        "can_reveal_title": bool(actor) and actor.get("role") == "admin",
        "status": row.get("status") or "processing",
        "device_model": row.get("device_model") or "PN0300",
        "source": row.get("source") or "plaud_note_pro",
        "duration_seconds": float(row.get("duration_seconds") or 0),
        "meeting_date": core.iso(row.get("meeting_date")),
        "received_at": core.iso(row.get("received_at")),
        "participants_count": len(row.get("participants") or []),
    }
    if detail:
        result.update({
            "participants": row.get("participants") or [],
            "summary": row.get("summary") or "",
            "decisions": row.get("decisions") or [],
            "action_items": row.get("action_items") or [],
            "transcript": row.get("transcript") or "",
            "source_url": row.get("source_url") or "",
        })
    return result


def _where_clause(query="", status=""):
    clauses = []
    params = []
    status = _clean_text(status, 20).lower()
    if status in {"completed", "processing", "failed"}:
        clauses.append("status=%s")
        params.append(status)
    query = _clean_text(query, 100)
    if query:
        clauses.append("title ILIKE %s")
        params.append(f"%{query}%")
    return (" WHERE " + " AND ".join(clauses) if clauses else ""), params


def list_meetings(actor, query="", status="", page=1, page_size=30, reveal_titles=False):
    page = _bounded_int(page, 1, 1, 100000)
    page_size = _bounded_int(page_size, 30, 5, 100)
    if core.DB_ENABLED:
        where, params = _where_clause(query, status)
        count = core.fetchone(f"SELECT count(*) AS count FROM plaud_device_meetings{where}", tuple(params)) or {"count": 0}
        rows = core.fetchall(
            f"""SELECT * FROM plaud_device_meetings{where}
                ORDER BY COALESCE(meeting_date,received_at) DESC LIMIT %s OFFSET %s""",
            tuple(params + [page_size, (page - 1) * page_size]),
        )
        total = int(count["count"])
    else:
        rows = list(_memory_meetings.values())
        if status:
            rows = [row for row in rows if row.get("status") == status]
        if query:
            rows = [row for row in rows if query.lower() in row.get("title", "").lower()]
        rows.sort(key=lambda row: row.get("meeting_date") or row.get("received_at"), reverse=True)
        total = len(rows)
        rows = rows[(page - 1) * page_size:page * page_size]
    return {
        "items": [public_meeting(row, actor, reveal_titles=reveal_titles) for row in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def stats():
    if core.DB_ENABLED:
        row = core.fetchone(
            """SELECT count(*) AS total,
                      count(*) FILTER (WHERE (COALESCE(meeting_date,received_at) AT TIME ZONE 'Asia/Seoul')::date = (now() AT TIME ZONE 'Asia/Seoul')::date) AS today,
                      count(*) FILTER (WHERE date_trunc('month',COALESCE(meeting_date,received_at) AT TIME ZONE 'Asia/Seoul') = date_trunc('month',now() AT TIME ZONE 'Asia/Seoul')) AS this_month,
                      count(*) FILTER (WHERE status='completed') AS completed
               FROM plaud_device_meetings"""
        ) or {}
        return {key: int(row.get(key) or 0) for key in ("total", "today", "this_month", "completed")}
    now = datetime.now(KST)
    values = list(_memory_meetings.values())
    dates = [(row.get("meeting_date") or row.get("received_at") or datetime.now(timezone.utc)).astimezone(KST) for row in values]
    return {
        "total": len(values),
        "today": sum(1 for value in dates if value.date() == now.date()),
        "this_month": sum(1 for value in dates if (value.year, value.month) == (now.year, now.month)),
        "completed": sum(1 for row in values if row.get("status") == "completed"),
    }


def _find_meeting(meeting_id):
    if core.DB_ENABLED:
        return core.fetchone("SELECT * FROM plaud_device_meetings WHERE id=%s", (str(meeting_id),))
    return _memory_meetings.get(str(meeting_id))


def get_meeting(meeting_id, actor, reveal_titles=False):
    row = _find_meeting(meeting_id)
    if not row:
        raise core.AppError("기기 회의록을 찾을 수 없습니다.", 404)
    return public_meeting(row, actor, detail=True, reveal_titles=reveal_titles)


def _join_lines(values):
    return "\n".join(f"• {value}" for value in (values or [])) or "-"


def build_excel(meeting_id, actor):
    if actor.get("role") != "admin":
        raise core.AppError("원문 제목이 포함된 회의록 Excel은 관리자만 다운로드할 수 있습니다.", 403)
    row = _find_meeting(meeting_id)
    if not row:
        raise core.AppError("기기 회의록을 찾을 수 없습니다.", 404)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "회의록"
    green = "0B6B57"
    light_green = "EAF5F1"
    border = Border(*( [Side(style="thin", color="D9E5E1")] * 4 ))
    sheet.merge_cells("A1:F2")
    sheet["A1"] = "메드파크 회의록"
    sheet["A1"].font = Font(name="맑은 고딕", size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=green)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    meeting_date = row.get("meeting_date") or row.get("received_at")
    if isinstance(meeting_date, datetime):
        meeting_date = meeting_date.astimezone(KST).strftime("%Y-%m-%d %H:%M")
    meta = [
        ("회의명", row.get("title") or "PLAUD 기기 회의록"),
        ("회의 일시", meeting_date or "-"),
        ("참석자", ", ".join(row.get("participants") or []) or "-"),
        ("녹음 기기", f"PLAUD Note Pro ({row.get('device_model') or 'PN0300'})"),
        ("연동 방식", "PLAUD Note Pro → Zapier → MedPark One"),
    ]
    current_row = 3
    for label, value in meta:
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
        sheet.cell(current_row, 1, label)
        sheet.cell(current_row, 2, value)
        current_row += 1

    sections = [
        ("회의 요약", row.get("summary") or "-"),
        ("주요 결정사항", _join_lines(row.get("decisions"))),
        ("후속 조치사항", _join_lines(row.get("action_items"))),
        ("전체 전사 내용", row.get("transcript") or "-"),
    ]
    for label, value in sections:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        heading = sheet.cell(current_row, 1, label)
        heading.font = Font(name="맑은 고딕", bold=True, color=green)
        heading.fill = PatternFill("solid", fgColor=light_green)
        current_row += 1
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=6)
        content = sheet.cell(current_row, 1, value)
        content.alignment = Alignment(vertical="top", wrap_text=True)
        current_row += 4

    for row_cells in sheet.iter_rows(min_row=3, max_row=current_row - 1, min_col=1, max_col=6):
        for cell in row_cells:
            cell.font = cell.font.copy(name="맑은 고딕", size=10)
            cell.border = border
            if cell.column == 1 and cell.row <= 7:
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color=green)
                cell.fill = PatternFill("solid", fgColor=light_green)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, 7):
        sheet.column_dimensions[get_column_letter(column)].width = 18 if column == 1 else 15
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    safe_title = re.sub(r"[^0-9A-Za-z가-힣._-]+", "_", row.get("title") or "회의록")[:80]
    return buffer, f"{safe_title}_회의록.xlsx"
